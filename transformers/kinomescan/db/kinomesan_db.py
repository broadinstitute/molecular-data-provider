
#from asyncio.windows_events import NULL
from re import I
from openpyxl import load_workbook
import urllib.request  
import pandas as pd

# Download datasets and merge into table
# Functions for building xlsx 
def build_activity_xlsx(dataset_file):
    '''
    Argument: HMS-LINCS_KinomeScan_Datasets_2018-01-18.xlsx file
    Output: Builds kinomescan_raw_data.xlsx (all activity data),
            db/data/kinomescan_activity.xlsx (filtered and arranged activity data)
    '''
    dataset_info = pd.read_excel(dataset_file)
    dataset_id = dataset_info["dataset_id"].tolist()
    df_list = []
    for id in dataset_id:
        id = str(id)
        url = 'https://lincs.hms.harvard.edu/db/datasets/'+id+'/results?search=&output_type=.xlsx'
        dataset_url = 'https://lincs.hms.harvard.edu/db/datasets/'+id+"/"
        result_file = "db/data/dataset_"+id+".xlsx"
        urllib.request.urlretrieve(url,result_file)  
        df = pd.read_excel(result_file)
        num_rows = len(df.index)
        df_dataset_id = [id]*num_rows
        df_dataset_url = [dataset_url]*num_rows
        df["dataset_id"] = df_dataset_id
        df["dataset_url"] = df_dataset_url
        df_list.append(df)
    activity_merged = pd.concat(df_list, ignore_index=True)
    activity_merged.to_excel('db/data/kinomescan_raw_data.xlsx', index=False)
    activity_df = pd.DataFrame(columns=["SM_HMS_LINCS_ID", "SALT_HMS_LINCS_ID","PROTEIN_HMS_LINCS_ID", "PERCENT_CONTROL", "KD", "ASSAY_COMPOUND_CONC","CONC_UNIT","DATASET_ID","DATASET_URL"])
    activity_index=0
    for index, row in activity_merged.iterrows():
        # eliminating batch number and salt id from HMSL id
        id = str(row[0])
        sm_id = "HMSL"+id[:-6]
        # properly formating salt id
        salt_id = id[6:9]+"-101"
        activity_df.loc[activity_index] = [sm_id,salt_id,row[2],row[4],row[9],row[5],row[6],row[7],row[8]]
        activity_index+=1
    activity_df.to_excel('db/data/kinomescan_activity.xlsx', index=False)

def build_protein_xlsx(activity_file, protein_file): 
    '''
    Argument: kinomescan_activity.xlsx file, proteins_20230203170917.xlsx file
    Output: kinomescan_protein.xlsx (filtered and arranged protein data)
    '''
    activity_info = pd.read_excel(activity_file)
    protein_info = pd.read_excel(protein_file)
    protein_id = activity_info["PROTEIN_HMS_LINCS_ID"].tolist()
    protein_id = list(set(protein_id))
    protein_info.drop(columns=["Alternative Names", "Provider", "Provider Catalog ID", "Protein Purity", "Protein Complex", "Isoform", "Protein Type", "Reference", "Comments", "Date Publicly Available", "Most Recent Update"], axis=1, inplace= True)
    protein_df = pd.DataFrame(columns= ["PROTEIN_NAME", "PROTEIN_HMS_LINCS_ID", "UNIPROT_ID", "AMINO_ACID_SEQUENCE","GENE_SYMBOL","GENE_ID","PROTEIN_SOURCE","MUTATION", "PHOSPHORYLATION_STATE", "DOMAIN","PROTEIN_DESCRIPTION", "SOURCE_ORGANISM"])
    protein_index = 0 
    for index, row in protein_info.iterrows():
        if row[1] in protein_id:
            protein_df.loc[protein_index] = [row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11]]
            protein_index+=1
    protein_df.to_excel('db/data/kinomescan_protein.xlsx', index=False)

def build_small_molecule_and_names_xlsx(activity_file, sm_file): 
    '''
    Argument: kinomescan_activity.xlsx file, small_molecule_20230120203823.xlsx file
    Output: kinomescan_small_molecule.xlsx (filtered and arranged small molecule data)
            kinomescan_small_molecule_names.xlsx (filtered and arranged small molecule names data)

    '''
    activity_info = pd.read_excel(activity_file)
    small_molecule_info = pd.read_excel(sm_file)
    hmsl_id = activity_info["SM_HMS_LINCS_ID"].tolist()
    hmsl_id = list(set(hmsl_id))
    sm_index=0
    sm_name_index=0
    small_molecule_df = pd.DataFrame(columns= ["SM_NAME","SM_LINCS_ID","SM_HMS_LINCS_ID","PUBCHEM_CID","CHEMBL_ID","MOLECULAR_MASS","INCHI","INCHI_KEY","SMILES"])
    small_molecule_name_df = pd.DataFrame(columns= ["SM_ALTERNATIVE_NAMES","SM_HMS_LINCS_ID"])
    for id in hmsl_id:
        id_number = str(id[4:])
        for index, row in small_molecule_info.iterrows():
            row_id = row[0][:-4]
            if row_id == id_number:
                if row[7] == "restricted":
                    row[7] = None
                    row[8] = None
                    row[9] = None
                    row[10] = None
                small_molecule_df.loc[sm_index] = [row[1], row[3], id, row[4], row[6], row[7], row[8], row[9], row[10]]
                sm_index+=1
                # Filtering out NaN
                if isinstance(row[2], str):
                    # Seperating multiple names into seprate rows
                    if ";" in row[2]:
                        sm_names = row[2].split("; ")
                        for name in sm_names:
                            small_molecule_name_df.loc[sm_name_index] = [name, id]
                            sm_name_index+=1
                    else:
                        small_molecule_name_df.loc[sm_name_index] = [row[2], id]
                        sm_name_index+=1
    small_molecule_df.drop_duplicates(inplace=True)
    small_molecule_name_df.drop_duplicates(inplace=True)
    small_molecule_df.to_excel('db/data/kinomescan_small_molecule.xlsx', index=False)
    small_molecule_name_df.to_excel('db/data/kinomescan_small_molecule_names.xlsx', index=False)

def build_salt_xlsx(salt_file): 
    '''
    Argument: small_molecule_20230210191036.xlsx file
    Output: kinomescan_salt.xlsx (filtered and arranged salt data)
    '''
    salt_info = pd.read_excel(salt_file)
    salt_info = salt_info.drop(columns=["Alternative Names", "LINCS ID", "PubChem CID", "ChEBI ID", "ChEMBL ID", "Relevant Citations", "Comments", "Most Recent Update", "Date Publicly Available"], axis = 1)
    for index, row in salt_info.iterrows():
        if row[3] == "-":
            salt_info.loc[index] = [row[0], row[1], None, None, None, None]
    salt_info.to_excel('db/data/kinomescan_salt.xlsx', index=False)

# Build sqlite data base
import sqlite3

connection = sqlite3.connect("db/data/KinomeScan.sqlite", check_same_thread=False)

SMALL_MOLECULE_TABLE = """
    CREATE TABLE SMALL_MOLECULE (
        SM_NAME TEXT, 
        SM_LINCS_ID TEXT,
        SM_HMS_LINCS_ID TEXT, 
        PUBCHEM_CID TEXT,
        CHEMBL_ID TEXT,
        MOLECULAR_MASS INT, 
        INCHI TEXT,
        INCHI_KEY TEXT,
        SMILES TEXT
    )
"""
SMALL_MOLECULE_NAMES_TABLE = """
    CREATE TABLE SMALL_MOLECULE_NAMES (
        SM_ALTERNATIVE_NAMES TEXT, 
        SM_HMS_LINCS_ID TEXT
    )
"""
SALT_TABLE = """
    CREATE TABLE SALT (
        SALT_HMS_LINCS_ID TEXT,
        SALT_NAME TEXT,
        MOLECULAR_MASS INT, 
        INCHI TEXT, 
        INCHI_KEY TEXT, 
        SMILES TEXT
    )
"""
PROTEIN_TABLE = """
    CREATE TABLE PROTEIN (
        PROTEIN_NAME TEXT, 
        PROTEIN_HMS_LINCS_ID TEXT, 
        UNIPROT_ID TEXT,
        AMINO_ACID_SEQUENCE TEXT,
        GENE_SYMBOL TEXT, 
        GENE_ID TEXT,
        PROTEIN_SOURCE TEXT,
        MUTATION TEXT, 
        PHOSPHORYLATION_STATE TEXT,
        DOMAIN TEXT,
        PROTEIN_DESCRIPTION TEXT,
        SOURCE_ORGANISM TEXT
    )
"""
ACTIVITY_TABLE = """
    CREATE TABLE ACTIVITY (
        SM_HMS_LINCS_ID TEXT, 
        SALT_HMS_LINCS_ID TEXT,
        PROTEIN_HMS_LINCS_ID TEXT, 
        PERCENT_CONTROL INT, 
        KD INT,
        ASSAY_COMPOUND_CONC INT, 
        CONC_UNIT TEXT, 
        DATASET_ID INT ,
        DATASET_URL TEXT
    )
"""
def create_tables():
    cur = connection.cursor()
    cur.execute(SMALL_MOLECULE_TABLE)
    cur.execute(SMALL_MOLECULE_NAMES_TABLE)
    cur.execute(SALT_TABLE)
    cur.execute(PROTEIN_TABLE)
    cur.execute(ACTIVITY_TABLE)
    cur.close()
    connection.commit()

def insert_small_molecule(cur, sm_name, sm_lincs_id, sm_hms_id, pubchem_cid, chembl_id, molecular_mass, inchi, inchi_key, smiles):
    statement = """
        INSERT INTO SMALL_MOLECULE (SM_NAME, SM_LINCS_ID, SM_HMS_LINCS_ID, PUBCHEM_CID, CHEMBL_ID, MOLECULAR_MASS, INCHI, INCHI_KEY, SMILES) VALUES (?,?,?,?,?,?,?,?,?)
    """
    cur.execute(statement,(sm_name, sm_lincs_id, sm_hms_id, pubchem_cid, chembl_id, molecular_mass, inchi, inchi_key, smiles))

def insert_small_molecule_names(cur, sm_alternative_names, sm_hms_id):
    statement = """
        INSERT INTO SMALL_MOLECULE_NAMES (SM_ALTERNATIVE_NAMES, SM_HMS_LINCS_ID) VALUES (?,?)
    """
    cur.execute(statement,(sm_alternative_names, sm_hms_id))

def insert_salt(cur, salt_hms_lincs_id, salt_name, molecular_mass, inchi, inchi_key, smiles):
    statement = """
        INSERT INTO SALT (SALT_HMS_LINCS_ID, SALT_NAME, MOLECULAR_MASS, INCHI, INCHI_KEY, SMILES) VALUES (?,?,?,?,?,?)
    """
    cur.execute(statement,(salt_hms_lincs_id, salt_name, molecular_mass, inchi, inchi_key, smiles))

def insert_protein(cur, protein_name, protein_hms_lincs_id, uniprot_id, amino_acid_sequence, gene_symbol, gene_id, protein_source, mutation, phosphorylation_state, domain, protein_description, source_organism):
    statement = """
        INSERT INTO PROTEIN (PROTEIN_NAME, PROTEIN_HMS_LINCS_ID, UNIPROT_ID, AMINO_ACID_SEQUENCE, GENE_SYMBOL, GENE_ID, PROTEIN_SOURCE, MUTATION, PHOSPHORYLATION_STATE, DOMAIN, PROTEIN_DESCRIPTION, SOURCE_ORGANISM) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """
    cur.execute(statement,(protein_name, protein_hms_lincs_id, uniprot_id, amino_acid_sequence, gene_symbol, gene_id, protein_source, mutation, phosphorylation_state, domain, protein_description, source_organism))

def insert_activity(cur, sm_hms_lincs_id, salt_hms_lincs_id, protein_hms_lincs_id, percent_control, kd, assay_compound_conc, conc_unit, dataset_id, dataset_url):
    statement = """
        INSERT INTO ACTIVITY (SM_HMS_LINCS_ID, SALT_HMS_LINCS_ID, PROTEIN_HMS_LINCS_ID, PERCENT_CONTROL, KD, ASSAY_COMPOUND_CONC, CONC_UNIT, DATASET_ID, DATASET_URL) VALUES (?,?,?,?,?,?,?,?,?)
    """
    cur.execute(statement,(sm_hms_lincs_id, salt_hms_lincs_id, protein_hms_lincs_id, percent_control, kd, assay_compound_conc, conc_unit, dataset_id, dataset_url))

def create_index(cur, table, column):
    statement = """
        CREATE INDEX {}_{}_IDX ON {}({});
    """
    cur.execute(statement.format(table, column, table, column))

def create_indexes():
    cur = connection.cursor()
    create_index(cur, 'SMALL_MOLECULE', 'SM_HMS_LINCS_ID')
    create_index(cur, 'SMALL_MOLECULE', 'SM_NAME')
    create_index(cur, 'SMALL_MOLECULE_NAMES', 'SM_ALTERNATIVE_NAMES')
    create_index(cur, "SALT", "SALT_HMS_LINCS_ID")
    create_index(cur, 'PROTEIN', 'PROTEIN_HMS_LINCS_ID')
    create_index(cur, 'PROTEIN', 'PROTEIN_NAME')
    create_index(cur, 'ACTIVITY', 'SM_HMS_LINCS_ID')
    create_index(cur, 'ACTIVITY', 'PROTEIN_HMS_LINCS_ID')
    create_index(cur, 'ACTIVITY', 'DATASET_ID')
    cur.close()
    connection.commit()

def load_small_molecule(filename):
    first_row = True
    cur = connection.cursor()
    wb = load_workbook(filename)
    wb_sheet = wb['Sheet1']
    for row in wb_sheet.iter_rows():
        if first_row:
            if len(row) != 9 or row[0].value != "SM_NAME" or row[1].value != "SM_LINCS_ID" or row[2].value != "SM_HMS_LINCS_ID" or \
            row[3].value != "PUBCHEM_CID" or row[4].value != "CHEMBL_ID" or row[5].value != "MOLECULAR_MASS" or row[6].value != "INCHI" or \
            row[7].value != "INCHI_KEY" or row[8].value != "SMILES":
                print('ERROR: wrong small molecule KinomeScan-file format')
                return
            first_row = False
        else:
            sm_name = row[0].value 
            sm_lincs_id= row[1].value
            sm_hms_id= row[2].value 
            pubchem_cid = row[3].value
            chembl_id= row[4].value
            molecular_mass= row[5].value 
            inchi= row[6].value
            inchi_key= row[7].value 
            smiles= row[8].value
            insert_small_molecule(cur, sm_name, sm_lincs_id, sm_hms_id, pubchem_cid, chembl_id, molecular_mass, inchi, inchi_key, smiles)
    cur.close()
    connection.commit()

def load_small_molecule_names(filename):
    first_row = True
    cur = connection.cursor()
    wb = load_workbook(filename)
    wb_sheet = wb['Sheet1']
    for row in wb_sheet.iter_rows():
        if first_row:
            if len(row) != 2 or row[0].value != "SM_ALTERNATIVE_NAMES" or row[1].value != "SM_HMS_LINCS_ID":
                print('ERROR: wrong small molecule names KinomeScan-file format')
                return
            first_row = False
        else:
            sm_alternative_names = row[0].value 
            sm_hms_id= row[1].value
            insert_small_molecule_names(cur, sm_alternative_names, sm_hms_id)
    cur.close()
    connection.commit()

def load_salt(filename):
    first_row = True
    cur = connection.cursor()
    wb = load_workbook(filename)
    wb_sheet = wb['Sheet1']
    for row in wb_sheet.iter_rows():
        if first_row:
            if len(row) != 6 or row[0].value != "HMS LINCS ID" or row[1].value != "Name" or \
            row[2].value != "Molecular Mass" or row[3].value != "InChi" or\
            row[4].value != "InChi Key" or row[5].value != "SMILES":
                print('ERROR: wrong salt KinomeScan-file format')
                return
            first_row = False
        else:
            salt_hms_lincs_id = row[0].value
            salt_name = row[1].value
            molecular_mass = row[2].value
            inchi = row[3].value
            inchi_key = row[4].value
            smiles = row[5].value
            insert_salt(cur, salt_hms_lincs_id, salt_name, molecular_mass, inchi, inchi_key, smiles)
    cur.close()
    connection.commit()

def load_protein(filename):
    first_row = True
    cur = connection.cursor()
    wb = load_workbook(filename)
    wb_sheet = wb['Sheet1']
    for row in wb_sheet.iter_rows():
        if first_row:
            if len(row) != 12 or row[0].value != "PROTEIN_NAME" or row[1].value != "PROTEIN_HMS_LINCS_ID" or row[2].value != "UNIPROT_ID" or \
            row[3].value != "AMINO_ACID_SEQUENCE" or row[4].value != "GENE_SYMBOL" or row[5].value != "GENE_ID" or row[6].value != "PROTEIN_SOURCE" or \
            row[7].value != "MUTATION" or row[8].value != "PHOSPHORYLATION_STATE" or row[9].value != "DOMAIN" or \
            row[10].value != "PROTEIN_DESCRIPTION" or row[11].value != "SOURCE_ORGANISM":
                print('ERROR: wrong protein KinomeScan-file format')
                return
            first_row = False
        else:
            protein_name = row[0].value 
            protein_hms_lincs_id= row[1].value
            uniprot_id= row[2].value 
            amino_acid_sequence = row[3].value
            gene_symbol= row[4].value
            gene_id= row[5].value 
            protein_source= row[6].value
            mutation= row[7].value 
            phosphorylation_state= row[8].value
            domain= row[9].value
            protein_description= row[10].value
            source_organism= row[11].value
            insert_protein(cur, protein_name, protein_hms_lincs_id, uniprot_id, amino_acid_sequence, gene_symbol, gene_id, protein_source, mutation, phosphorylation_state, domain, protein_description, source_organism)
    cur.close()
    connection.commit()

def load_activity(filename):
    first_row = True
    cur = connection.cursor()
    wb = load_workbook(filename)
    wb_sheet = wb['Sheet1']
    for row in wb_sheet.iter_rows():
        if first_row:
            if len(row) != 9 or row[0].value != "SM_HMS_LINCS_ID" or row[1].value != "SALT_HMS_LINCS_ID" or row[2].value != "PROTEIN_HMS_LINCS_ID" \
            or row[3].value != "PERCENT_CONTROL" or row[4].value != "KD" or row[5].value != "ASSAY_COMPOUND_CONC" or row[6].value != "CONC_UNIT" or row[7].value != "DATASET_ID" \
            or row[8].value != "DATASET_URL":
                print('ERROR: wrong activity KinomeScan-file format')
                return
            first_row = False
        else:
            sm_hms_lincs_id = row[0].value 
            salt_hms_lincs_id= row[1].value
            protein_hms_lincs_id= row[2].value 
            percent_control = row[3].value
            kd = row[4].value
            assay_compound_conc= row[5].value
            conc_unit= row[6].value 
            dataset_id= row[7].value
            dataset_url= row[8].value 
            insert_activity(cur, sm_hms_lincs_id, salt_hms_lincs_id, protein_hms_lincs_id, percent_control, kd, assay_compound_conc, conc_unit, dataset_id, dataset_url)
    cur.close()
    connection.commit()


def main():
    # build_activity_xlsx("db/data/HMS-LINCS_KinomeScan_Datasets_2018-01-18.xlsx")
    build_protein_xlsx('db/data/kinomescan_activity.xlsx', "db/data/proteins_20230203170917.xlsx")
    build_small_molecule_and_names_xlsx('db/data/kinomescan_activity.xlsx', "db/data/small_molecule_20230120203823.xlsx")
    build_salt_xlsx("db/data/small_molecule_20230210191036.xlsx")
    create_tables()
    load_small_molecule("db/data/kinomescan_small_molecule.xlsx")
    load_small_molecule_names("db/data/kinomescan_small_molecule_names.xlsx")
    load_salt("db/data/kinomescan_salt.xlsx")
    load_protein("db/data/kinomescan_protein.xlsx")
    load_activity("db/data/kinomescan_activity.xlsx")
    create_indexes()


if __name__ == '__main__':
    main()
            
